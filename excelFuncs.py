import openpyxl as pxl
import pandas as pd
from itertools import groupby
import sys
from finalAlg import maxClusterArea, actualClusterArea
from geoThings import newCircle
from convPolyFuncs import convPoly




filepath = r'D:\Work\Excel Files\Kajiado Data v1.xlsx'
krcp_filepath = r"D:\Work\Excel Files\KRCPLivestockdata Locations UTM Fixed Radii.xlsx"
pdtesting_filepath = r"D:\Work\Excel Files\KRCP Testing.csv"
krcp_resultpath = r"D:\Work\Excel Files\KRCPLivestockdata Locations UTM Fixed Radii Results.csv"


pdkrcp = pd.read_excel(krcp_filepath, sheet_name = 0)
pdtesting = pd.read_excel(krcp_filepath, sheet_name = 1)

# testinggroups = pdtesting.groupby(
#     ["Block","Zone","Veggie"]
# )

# def test1():
#     results = []
#     # print("test group:\n", testinggroups.groups)
#     for name, group in testinggroups:
#         group["Result"] = maxClusterArea(list(group["Radius"].values))
#         results.append(group["Result"])
#     return results

# df.apply(lambda x: f(x.col1, x.col2), axis = 1)

def getActualArea(accuracyCol, xcoordCol, ycoordCol, radiusCol):
    def go(n):
        if abs(n) > 150:
            return True
        else:
            return False
    
    if accuracyCol.apply(go).any() or radiusCol.apply(lambda x: x == 0).any():
        return None

    

def getAreasOfActualClusters():
    def go(n):
        if abs(n) > 150:
            return True
        else:
            return False
    
    results = []
    
    krcp_groups = pdkrcp.groupby(
        [pdkrcp["Date"].dt.date, "Zone", "Grazing_Block"],
        sort = False
    )
    print("num of groups:", len(krcp_groups))
    groupNum = 0
    for _, group in krcp_groups:
        groupNum += 1
        print("\n\n\n\n\n")
        print(groupNum,"\n")

        if group["GPS_Accuracy"].apply(go).any() or group["fixed_radius"].apply(lambda x: float(x) == 0.0).any():
            group["actual_area"] = None
            results.append(group["actual_area"])
        
        else:
            circleList = []
            for i in range(len(group)):
                print(group["POINT_X"].to_list()[i])
                print(group["POINT_Y"].to_list()[i])
                print(group["fixed_radius"].to_list()[i])
                
                circleList.append(
                    newCircle(
                        group["POINT_X"].to_list()[i],
                        group["POINT_Y"].to_list()[i],
                        group["fixed_radius"].to_list()[i]
                    )
                )
            
            try:
                print("doing area")
                group["actual_area"] = actualClusterArea(circleList)
                print("finished area")
            except:
                print("shit")
                group["actual_area"] = 0
            results.append(group["actual_area"])

        print(group[["Date", "Zone", "Grazing_Block", "GPS_Accuracy","POINT_X","POINT_Y","fixed_radius","actual_area"]])
        print("finished group", groupNum)

    
    pdkrcp["actual_area"] = pd.concat(results)
    print("\n\n\n\n\nResults:", pdkrcp[["GPS_Accuracy","POINT_X","POINT_Y","fixed_radius","actual_area"]])

    pdkrcp.to_csv(krcp_resultpath)


    


getAreasOfActualClusters()    


def applyMaxClusterArea():
    '''
    iterate over radius column
    for each cell, associate to it
        1. the date, block, and zone cells in the same row
        2. the result cell in the "maximum cluster area" column
    '''
    
    #openpyxl loading
    kaj = pxl.load_workbook(filepath)
    kaj1 = kaj[kaj.sheetnames[0]]
    
    #for each column, makes a dictionary entry of the value of the 1st row, and the rest of the cells
    def getCols(sheet): 
        cols = {}
        for col in sheet.iter_cols():
            title = col[0].value #whatever type col[0].value is, probably a string
            cols[title] = col[1:] #the non-first elements of the column, skipping the title
        return cols
    
    columns = getCols(kaj1)

    def assocCells(colName, dataCols, resultCol):
        #string, [string], string
        results = []
        for cell in columns[colName]:
            rowNum = cell.row - 2
            results.append([cell,[columns[name][rowNum].value for name in dataCols], columns[resultCol][rowNum]])
        return results

    stuff = assocCells("fixed radius", ["Date","Zone","Grazing Block"],"max cluster area")

    groupedRows = [list(grp) for k, grp in
        groupby(stuff, lambda c: c[1])]

    print("There are",len(groupedRows),"groups")

    def doit():
        for g in groupedRows:
            try:
                radii = [el[0].value for el in g]
                if None in radii:
                    mcArea = None
                elif 0 in radii:
                    nonzeroRadii = [r for r in radii if r != 0]
                    if nonzeroRadii == []:
                        mcArea = None
                    else:
                        mcArea = maxClusterArea(nonzeroRadii)
                else:
                    mcArea = maxClusterArea(radii)

                for cell in [el[2] for el in g]:
                    cell.value = mcArea

                print("Finished group")
            except:
                print(g)
                sys.exit(0)
        print("Finished all the groups")

    def saveit():
        resultpath = r"D:\Work\Excel Files\Kajiado Data Results.xlsx"
        kaj.save(resultpath)

    doit()
    saveit()
