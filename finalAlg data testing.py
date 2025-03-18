from geoThings import Circle, randomList
from packingAlgs import radSumAlg, polyAreaAlg, randomAlg
from convPolyFuncs import convPoly
from drawFuncsV2 import setup, pshow, pdraw, drawCircles, drawPolygon

from shapely import Polygon
import openpyxl as pxl
import timeit

def main():
    cells = ['B1','C1','D1','E1','F1','H1','I1','J1','K1','L1','N1','O1','P1','Q1','R1']
    cells2 = ['B1']
    cells3 = ['B1','C1','D1','E1','F1']

    radii = randomList(15,10,100)
    testc = radSumAlg(radii)
    testp = convPoly(testc)

    setup()
    drawCircles(testc)
    drawPolygon(testp)
    pshow()


def maxClusterArea(radii):
    if len(radii) <= 10:
        return convPoly(polyAreaAlg(radii)).area
    else:
        return convPoly(radSumAlg(radii)).area

#Excel
wb=pxl.load_workbook('Circle Data.xlsx')
cData=wb['Circle Data']

def getRadii():
    startCell=str(input('Enter starting cell: '))
    column=startCell[0]
    index=int(startCell[1])
    radii=[]
    while True:
        if cData[column+str(index)].value == None:
            break
        radii.append(
            float(cData[column+str(index)].value)
            )
        index+=1
    return radii

def directGetRadii(startCell):
    column=startCell[0]
    index=int(startCell[1])
    radii=[]
    while True:
        if cData[column+str(index)].value == None:
            break
        radii.append(
            float(cData[column+str(index)].value)
            )
        index+=1
    return radii

#For testing optimality
#takes circle data and returns data, nothing excel related
def optimalTest(iters, alg, radii):
    maxc = alg(radii)
    maxc_poly = convPoly(maxc)
    maxc_area = maxc_poly.area

    counterexampleCount = 0
    counterexampleAreaPercents = []

    for i in range(iters):
        randc = randomAlg(radii)
        randc_poly = convPoly(randc)
        randc_area = randc_poly.area

        if randc_area < maxc_area:
            counterexampleCount+=1
            counterexampleAreaPercents.append(randc_area/maxc_area*100)
        print('Finished iteration', i+1)

    numberOfCircles = len(radii)
    rangeOfRadii = max(radii)-min(radii)
    if counterexampleAreaPercents == []:
        avgCounterAreaPercent = 100
    else:
        avgCounterAreaPercent = sum(counterexampleAreaPercents)/len(counterexampleAreaPercents)

    return numberOfCircles, rangeOfRadii, iters, counterexampleCount, avgCounterAreaPercent

#takes a cell, gets the data from its column, puts it into the corresponding stuff in the excel file
def colAnalyze(iters, alg, cell):
    column = cell[0]

    radii = directGetRadii(cell)

    results = [column] + [v for v in optimalTest(iters, alg, radii)]

    if alg == radSumAlg:
        resultSheet = wb[wb.sheetnames[1]]
    if alg == polyAreaAlg:
        resultSheet = wb[wb.sheetnames[2]]

    for index, value in enumerate(results):
        row = index+1
        resultCell = str(column) + str(row)
        resultSheet[resultCell] = value

    wb.save('D:\Work\Python Scripts\Circle Data.xlsx')

if __name__ == '__main__':
    main()
